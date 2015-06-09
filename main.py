
from openpyxl import load_workbook

wb = load_workbook(filename = 'Test excel.xlsx')  
sheet1 = wb['Sheet1']
file_name = 'Test excel.xlsx'

sheet2 = wb.create_sheet(title="copy")
rowcount = 0
colcount = 0

last = 'A1'  
for row in sheet1.rows:
    rowcount += 1
    colcount = 0
    print (rowcount)
    for col in row:
        colcount += 1
        print (colcount)
        sheet2.cell(column=colcount, row=rowcount).value = col.value
  
wb.save(filename = file_name)


